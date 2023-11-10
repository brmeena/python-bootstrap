import sys
import traceback

sys.path.append("../")
from helpers import filehelper
import csv
import openpyxl
def readDataFromSheet(sheet_file):
    file_extension=filehelper.getFileExtension(sheet_file)
    if file_extension.lower()==".csv":
        return readCSVData(sheet_file)
    if file_extension.lower() in [".xls",".xlsx"]:
        return readExcelData(sheet_file)
    return None

def readCSVData(sheet_file):
    sheet_data=[]
    with open(sheet_file) as fi:
        csvreader=csv.reader(fi)
        for row in csvreader:
            sheet_data.append(row)
    return sheet_data

def readExcelData(sheet_file):
    sheet_data=[]
    dataframe = openpyxl.load_workbook(sheet_file)
    active_dataframe = dataframe.active
    for row in range(0, active_dataframe.max_row):
        row_item=[]
        for col in active_dataframe.iter_cols(1, active_dataframe.max_column):
            #print(col[row].value)
            row_item.append(str(col[row].value).rstrip("0."))
        sheet_data.append(row_item)
    return sheet_data

